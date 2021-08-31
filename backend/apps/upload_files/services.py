from io import BytesIO

from openpyxl import load_workbook

from .constants import WORKBOOK_COLUMNS, FORMAT_DATETIME


def get_value(ws, col_index):
    # Парсинг колонки
    for x in range(2, ws.max_row + 1):
        yield ws.cell(row=x, column=col_index).internal_value


def process_excel_file(file):
    wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
    columns = {}
    for sheet_name in wb.sheetnames:
        # формирование словаря с ключами before и after,
        # с заполнением данными из соответствующих им строк
        headers = {
            c.value: list(get_value(wb[sheet_name], c.col_idx))
            for c in next(wb[sheet_name].iter_rows(min_row=1, max_row=1))
            if c.value in WORKBOOK_COLUMNS
        }
        if headers:
            columns = headers
            break

    # получение списка уникальных значений после вычитания значения строк
    res = list(set([x - y for x, y in zip(columns.get('after'), columns.get('before'))]))

    if not res or len(res) > 1:
        return 'error'

    if res[0] > 0:
        status = f'added: {res[0]}'
    elif res[0] < 0:
        status = f'removed: {abs(res[0])}'
    else:
        status = 'nothing'
    return status


def format_datetime(date):
    # Приводит дату в заданный формат для вывода
    if date:
        return date.strftime(FORMAT_DATETIME)
    return None
